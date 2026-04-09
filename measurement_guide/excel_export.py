"""Build Measurement Guide Excel workbook (Cerave technical sheet + overview sheet)."""

import io
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from measurement_guide.datalayer_builders import build_measurement_script_excel
from measurement_guide.models import EXPORT_FORMAT_VERSION

MIN_ROW_PT = 18.0
_MAX_SCREENSHOT_PX = 200

# Empty rows in "Additional reference" for manual key/value notes in Excel.
_GUIDE_DICT_EMPTY_ROWS = 18


def _px_to_row_height_pt(px: float) -> float:
    return max(12.0, min(409.0, px * 72.0 / 96.0))


def _style_header(cell, *, fill, font, align, border):
    cell.fill = fill
    cell.font = font
    cell.alignment = align
    cell.border = border


def _fill_guide_overview_sheet(
    ws,
    events: List[Dict[str, Any]],
    reference_pairs: Optional[List[Tuple[str, str]]],
) -> None:
    """Populate first sheet: title, export info, per-event summary, optional KV table."""
    thin = Side(style="thin", color="FFB4B4B4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="FF1F4E79", end_color="FF1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=14, color="FF1F4E79")
    header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
    body_align = Alignment(wrap_text=True, vertical="top", horizontal="left")

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "Measurement guide — context & summary"
    t.font = title_font
    t.alignment = Alignment(vertical="center", horizontal="left")

    ws["A3"] = "Generated (UTC)"
    ws["B3"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    ws["A4"] = "Guide export format"
    ws["B4"] = EXPORT_FORMAT_VERSION

    for addr in ("A3", "A4"):
        ws[addr].font = Font(bold=True)

    start_row = 6
    ws.cell(row=start_row, column=1, value="Events summary")
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)

    hdr_row = start_row + 1
    summary_headers = [
        "#",
        "Page / URL",
        "Environment",
        "Notes",
        "GA4 event (recommended)",
        "Element type",
    ]
    for col, h in enumerate(summary_headers, 1):
        c = ws.cell(row=hdr_row, column=col, value=h)
        _style_header(c, fill=header_fill, font=header_font, align=header_align, border=border)

    data_row = hdr_row + 1
    for i, ev in enumerate(events, start=1):
        r = data_row + i - 1
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=ev.get("page_url") or "")
        ws.cell(row=r, column=3, value=ev.get("environment") or "")
        ws.cell(row=r, column=4, value=ev.get("notes") or "")
        ws.cell(row=r, column=5, value=ev.get("ga4_event_name") or "")
        ws.cell(row=r, column=6, value=ev.get("type") or "")
        for c in range(1, 7):
            ws.cell(row=r, column=c).alignment = body_align
            ws.cell(row=r, column=c).border = border

    dict_title_row = data_row + len(events) + 1
    ws.cell(row=dict_title_row, column=1, value="Additional reference (key / value)")
    ws.cell(row=dict_title_row, column=1).font = Font(bold=True, size=12)

    dhdr = dict_title_row + 1
    for col, h in enumerate(["Key", "Value"], 1):
        c = ws.cell(row=dhdr, column=col, value=h)
        _style_header(c, fill=header_fill, font=header_font, align=header_align, border=border)

    pairs = list(reference_pairs or [])
    for j in range(max(_GUIDE_DICT_EMPTY_ROWS, len(pairs))):
        r = dhdr + 1 + j
        k, v = ("", "")
        if j < len(pairs):
            k, v = pairs[j][0], pairs[j][1]
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
        for c in (1, 2):
            ws.cell(row=r, column=c).alignment = body_align
            ws.cell(row=r, column=c).border = border

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 14


def build_measurement_workbook(
    events: List[Dict[str, Any]],
    reference_pairs: Optional[List[Tuple[str, str]]] = None,
) -> io.BytesIO:
    """
    Produce .xlsx: sheet 'Guide overview' (context + summary + reference dict),
    sheet 'Measurement Guide' (Cerave-style 5 columns: screenshot, how, script, var, values).
    """
    headers = [
        "Screenshot",
        "how it is triggered",
        "Script",
        "Variable",
        "Values",
    ]

    _script_font = Font(name="Consolas", size=10)
    _thin = Side(style="thin", color="FFB4B4B4")
    _thin_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _header_fill = PatternFill(start_color="FF1F4E79", end_color="FF1F4E79", fill_type="solid")
    _header_font = Font(bold=True, color="FFFFFF", size=11)
    _header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
    _body_align = Alignment(wrap_text=True, vertical="top", horizontal="left")

    wb = Workbook()
    ws = wb.active
    ws.title = "Measurement Guide"
    ws.row_dimensions[1].height = 22

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _header_fill
        cell.font = _header_font
        cell.alignment = _header_align
        cell.border = _thin_border

    row_num = 2
    for ev in events:
        how = ev.get("how") or ""
        script = build_measurement_script_excel(ev.get("datalayer") or {})
        screenshot_bytes = ev.get("screenshot")
        pairs = list((ev.get("datalayer") or {}).items())
        n_pairs = len(pairs)

        block_start = row_num
        block_end = block_start + max(0, n_pairs)

        ws.cell(row=block_start, column=2, value=how)
        ws.cell(row=block_start, column=3, value=script)

        first_var_row = block_start + 1
        for i, (var, val) in enumerate(pairs):
            r = first_var_row + i
            ws.cell(row=r, column=4, value=var)
            ws.cell(row=r, column=5, value=str(val))
            for c in (4, 5):
                ws.cell(row=r, column=c).alignment = _body_align
                ws.cell(row=r, column=c).border = _thin_border

        for r in range(block_start, block_end + 1):
            for c in (4, 5):
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    cell.border = _thin_border
                    cell.alignment = _body_align

        ws.merge_cells(
            start_row=block_start,
            start_column=1,
            end_row=block_end,
            end_column=1,
        )
        ws.merge_cells(
            start_row=block_start,
            start_column=2,
            end_row=block_end,
            end_column=2,
        )
        ws.merge_cells(
            start_row=block_start,
            start_column=3,
            end_row=block_end,
            end_column=3,
        )

        for col in (1, 2, 3):
            mc = ws.cell(row=block_start, column=col)
            mc.alignment = _body_align
            mc.border = _thin_border
            if col == 3:
                mc.font = _script_font

        num_merged_rows = block_end - block_start + 1

        if screenshot_bytes:
            try:
                img = XLImage(io.BytesIO(screenshot_bytes))
                if img.width > _MAX_SCREENSHOT_PX:
                    ratio = _MAX_SCREENSHOT_PX / img.width
                    img.width = int(img.width * ratio)
                    img.height = int(img.height * ratio)
                img_height_pt_total = _px_to_row_height_pt(float(img.height)) + 8.0
                even_split = img_height_pt_total / float(num_merged_rows)
                per_row_h = max(MIN_ROW_PT, even_split)
                for r in range(block_start, block_end + 1):
                    ws.row_dimensions[r].height = per_row_h
                ws.add_image(img, f"A{block_start}")
            except Exception:
                ws.cell(row=block_start, column=1, value="[Invalid image]")
                for r in range(block_start, block_end + 1):
                    ws.row_dimensions[r].height = max(
                        MIN_ROW_PT, 60.0 / float(max(1, num_merged_rows))
                    )
        else:
            for r in range(block_start, block_end + 1):
                ws.row_dimensions[r].height = max(
                    MIN_ROW_PT, 80.0 / float(max(1, num_merged_rows))
                )

        row_num = block_end + 2

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 58
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 32

    overview = wb.create_sheet("Guide overview", 0)
    _fill_guide_overview_sheet(overview, events, reference_pairs)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
