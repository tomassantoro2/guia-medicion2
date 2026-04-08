import streamlit as st
import io
import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage


def _escape_js_single(s: str) -> str:
    return str(s).replace("\\", "\\\\").replace("'", "\\'")


def build_measurement_script_excel(dl: dict) -> str:
    """Multi-line <script> + dataLayer.push for Excel (single-quoted, readable layout)."""
    if not dl:
        return "<script>\ndataLayer.push({});\n</script>"
    lines = ["<script>", "dataLayer.push({"]
    items = list(dl.items())
    for i, (k, v) in enumerate(items):
        comma = "," if i < len(items) - 1 else ""
        lines.append(f"  '{_escape_js_single(k)}': '{_escape_js_single(v)}'{comma}")
    lines.append("});")
    lines.append("</script>")
    return "\n".join(lines)


def _safe_js_double_quoted(v) -> str:
    return str(v).replace("\\", "\\\\").replace('"', '\\"')


def build_measurement_script_preview(dl: dict) -> str:
    """Pretty dataLayer.push for Streamlit preview (double-quoted, no script tags)."""
    if not dl:
        return "dataLayer.push({});"
    lines = ["dataLayer.push({"]
    items = list(dl.items())
    for i, (k, v) in enumerate(items):
        comma = "," if i < len(items) - 1 else ""
        lines.append(f'  {k}: "{_safe_js_double_quoted(v)}"{comma}')
    lines.append("});")
    return "\n".join(lines)


st.set_page_config(page_title="Measurement Guide – Phase 1", layout="centered")

# Load MTP (Beautify PE) events from JSON when present
MTP_EVENTS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mtp_events.json")
MTP_EVENTS = []
if os.path.isfile(MTP_EVENTS_PATH):
    try:
        with open(MTP_EVENTS_PATH, encoding="utf-8") as f:
            MTP_EVENTS = json.load(f)
    except Exception:
        MTP_EVENTS = []

if "events" not in st.session_state:
    st.session_state.events = []

if "extra_params" not in st.session_state:
    st.session_state.extra_params = []

st.title("📏 Measurement Guide Generator")
st.caption("Phase 1 · Custom events · dataLayer.push")

event_source = st.radio(
    "Event source",
    ["Custom event", "MTP event (Beautify PE)"] if MTP_EVENTS else ["Custom event"],
    key="event_source",
    horizontal=True
)

event_type = st.selectbox(
    "Event type",
    ["Button", "Banner", "Link"],
    key="event_type"
)

how_triggered = st.text_area(
    "How it is triggered",
    placeholder="e.g. When the user clicks the Hot Sale button on the home page",
    key="how_triggered"
)

screenshot_file = st.file_uploader(
    "📷 Screenshot",
    type=["png", "jpg", "jpeg"],
    key="screenshot_upload",
    help="Upload an image to include it in the Screenshot column of the Excel export."
)
if screenshot_file is not None:
    st.image(screenshot_file, caption="Preview", use_container_width=False, width=200)

using_mtp = event_source == "MTP event (Beautify PE)" and MTP_EVENTS

if using_mtp:
    mtp_options = [f"{e['description']} — {e['event_name']}" for e in MTP_EVENTS]
    mtp_selected_idx = st.selectbox(
        "MTP event",
        range(len(MTP_EVENTS)),
        format_func=lambda i: mtp_options[i],
        key="mtp_event_idx"
    )
    mtp_event = MTP_EVENTS[mtp_selected_idx]
    dl = dict(mtp_event["dl"])
    st.caption(
        f"Selected event: **{mtp_event['event_name']}**. The dataLayer is filled in automatically."
    )
else:
    event_base = st.selectbox(
        "GTM event (`event`)",
        ["uaevent", "nievent", "socialInt", "Custom"],
        key="event_base"
    )

    if event_base == "Custom":
        event_value = st.text_input(
            "Custom event name",
            placeholder="e.g. my_custom_event",
            key="event_custom"
        )
    else:
        event_value = event_base

    event_name = st.text_input(
        "event_name (suggested)",
        placeholder="e.g. button_click",
        key="event_name"
    )

    st.markdown("### Suggested parameters")
    st.caption(
        "Optional fields: leave blank to omit them from the dataLayer. Only non-empty values are included."
    )

    eventCategory = st.text_input(
        "eventCategory",
        placeholder="e.g. interaction",
        key="eventCategory"
    )

    eventAction = st.text_input(
        "eventAction",
        placeholder="e.g. click",
        key="eventAction"
    )

    eventLabel = st.text_input(
        "eventLabel",
        placeholder="Optional; omitted if left blank",
        key="eventLabel"
    )

    st.markdown("### Additional parameters")
    st.caption("Add as many as required. Enter name and value, then click Add.")

    col_extra1, col_extra2, col_extra3 = st.columns([2, 2, 1])
    with col_extra1:
        extra_key = st.text_input(
            "Name",
            key="extra_key",
            label_visibility="collapsed",
            placeholder="Parameter name",
        )
    with col_extra2:
        extra_value = st.text_input(
            "Value",
            key="extra_value",
            label_visibility="collapsed",
            placeholder="Parameter value",
        )
    with col_extra3:
        st.write("")
        if st.button("➕ Add", key="add_extra_param"):
            if extra_key and extra_value:
                st.session_state.extra_params.append({"key": extra_key, "value": extra_value})
                if "extra_key" in st.session_state:
                    del st.session_state["extra_key"]
                if "extra_value" in st.session_state:
                    del st.session_state["extra_value"]
                st.rerun()

    if st.session_state.extra_params:
        st.caption("Added parameters:")
        for i, p in enumerate(st.session_state.extra_params):
            col1, col2, col3 = st.columns([2, 2, 1])
            with col1:
                st.text(p["key"])
            with col2:
                st.text(p["value"])
            with col3:
                if st.button("🗑️", key=f"del_param_{i}"):
                    st.session_state.extra_params.pop(i)
                    st.rerun()

    dl = {}

    if event_value:
        dl["event"] = str(event_value)

    if event_name:
        dl["event_name"] = str(event_name)

    _ec = (eventCategory or "").strip()
    if _ec:
        dl["eventCategory"] = _ec

    _ea = (eventAction or "").strip()
    if _ea:
        dl["eventAction"] = _ea

    _el = (eventLabel or "").strip()
    if _el:
        dl["eventLabel"] = _el

    for p in list(st.session_state.extra_params):
        dl[str(p["key"])] = str(p["value"])

st.markdown("### 📜 dataLayer.push preview")

st.code(build_measurement_script_preview(dl), language="javascript")

if st.button("➕ Add event to guide", type="primary"):
    if using_mtp:
        valid = bool(dl)
        err_msg = "There is no dataLayer for the selected MTP event." if not valid else None
    else:
        ev_base = st.session_state.get("event_base", "")
        ev_custom = st.session_state.get("event_custom", "")
        ev_name = st.session_state.get("event_name", "")
        event_value = ev_custom if ev_base == "Custom" else ev_base
        valid = bool(event_value) and bool(ev_name)
        if ev_base == "Custom" and not ev_custom:
            valid = False
        if not valid:
            if not event_value:
                err_msg = "Please define the event (uaevent, custom, etc.)."
            elif not ev_name:
                err_msg = "Please define an event_name."
            elif ev_base == "Custom":
                err_msg = "When Custom is selected, please enter the custom event name."
            else:
                err_msg = "Please complete the required fields."
        else:
            err_msg = None

    if not valid:
        st.error(err_msg)
    else:
        screenshot_bytes = screenshot_file.read() if screenshot_file else None
        st.session_state.events.append({
            "type": event_type,
            "how": how_triggered,
            "datalayer": dl,
            "screenshot": screenshot_bytes
        })
        st.success("Event added to the guide")

        keys_to_clear = [
            "event_type", "how_triggered", "event_base", "event_custom",
            "event_name", "eventCategory", "eventAction", "eventLabel",
            "extra_key", "extra_value", "screenshot_upload", "mtp_event_idx"
        ]
        for key in keys_to_clear:
            if key in st.session_state:
                del st.session_state[key]
        st.rerun()

st.divider()
st.subheader("📄 Events in guide")

if not st.session_state.events:
    st.info("No events have been added yet")
else:
    # Cerave-style Excel layout: Screenshot | how it is triggered | Script | Variable | Values
    headers = ["Screenshot", "how it is triggered", "Script", "Variable", "Values"]

    # Excel styling
    MIN_ROW_PT = 18.0
    _script_font = Font(name="Consolas", size=10)
    _thin = Side(style="thin", color="FFB4B4B4")
    _thin_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _header_fill = PatternFill(start_color="FF1F4E79", end_color="FF1F4E79", fill_type="solid")
    _header_font = Font(bold=True, color="FFFFFF", size=11)
    _header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
    _body_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
    _max_screenshot_px = 200

    def _px_to_row_height_pt(px: float) -> float:
        return max(12.0, min(409.0, px * 72.0 / 96.0))

    wb = Workbook()
    ws = wb.active
    ws.title = "Measurement Guide"
    ws.row_dimensions[1].height = 22

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _header_fill
        cell.font = _header_font
        cell.alignment = _header_align
        cell.border = _thin_border

    row_num = 2
    for ev in st.session_state.events:
        how = ev["how"] or ""
        script = build_measurement_script_excel(ev["datalayer"])
        screenshot_bytes = ev.get("screenshot")
        pairs = list(ev["datalayer"].items())
        n_pairs = len(pairs)

        block_start = row_num
        # One row for how/script; variables start on the next row (no extra blank row in D/E).
        # block_end is the last row of this event (last Variable/Values row when n_pairs > 0).
        block_end = block_start + max(0, n_pairs)

        ws.cell(row=block_start, column=2, value=how)
        ws.cell(row=block_start, column=3, value=script)

        first_var_row = block_start + 1
        for i, (var, val) in enumerate(pairs):
            r = first_var_row + i
            ws.cell(row=r, column=4, value=var)
            ws.cell(row=r, column=5, value=str(val))
            for c in (4, 5):
                ws.cell(row=r, column=c).alignment = _body_align
                ws.cell(row=r, column=c).border = _thin_border

        for r in range(block_start, block_end + 1):
            for c in (4, 5):
                cell = ws.cell(row=r, column=c)
                if cell.value is None:
                    cell.border = _thin_border
                    cell.alignment = _body_align

        ws.merge_cells(
            start_row=block_start,
            start_column=1,
            end_row=block_end,
            end_column=1,
        )
        ws.merge_cells(
            start_row=block_start,
            start_column=2,
            end_row=block_end,
            end_column=2,
        )
        ws.merge_cells(
            start_row=block_start,
            start_column=3,
            end_row=block_end,
            end_column=3,
        )

        for col in (1, 2, 3):
            mc = ws.cell(row=block_start, column=col)
            mc.alignment = _body_align
            mc.border = _thin_border
            if col == 3:
                mc.font = _script_font

        num_merged_rows = block_end - block_start + 1

        if screenshot_bytes:
            try:
                img = XLImage(io.BytesIO(screenshot_bytes))
                if img.width > _max_screenshot_px:
                    ratio = _max_screenshot_px / img.width
                    img.width = int(img.width * ratio)
                    img.height = int(img.height * ratio)
                img_height_pt_total = _px_to_row_height_pt(float(img.height)) + 8.0
                even_split = img_height_pt_total / float(num_merged_rows)
                per_row_h = max(MIN_ROW_PT, even_split)
                for r in range(block_start, block_end + 1):
                    ws.row_dimensions[r].height = per_row_h
                ws.add_image(img, f"A{block_start}")
            except Exception:
                ws.cell(row=block_start, column=1, value="[Invalid image]")
                for r in range(block_start, block_end + 1):
                    ws.row_dimensions[r].height = max(
                        MIN_ROW_PT, 60.0 / float(max(1, num_merged_rows))
                    )
        else:
            for r in range(block_start, block_end + 1):
                ws.row_dimensions[r].height = max(
                    MIN_ROW_PT, 80.0 / float(max(1, num_merged_rows))
                )

        row_num = block_end + 2

    # Column widths: A sized for screenshot; script column wide for code
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 78
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 38

    # Save to buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # Download button
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"guia_medicion_{ts}.xlsx"

    st.download_button(
        label="📥 Download Excel (Cerave-style layout)",
        data=excel_buffer,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_excel"
    )

    st.write("")

    for i, ev in enumerate(st.session_state.events, start=1):
        with st.expander(f"Event {i}"):
            if ev.get("screenshot"):
                st.image(io.BytesIO(ev["screenshot"]), caption="Screenshot", width=200)
            st.write("**How it is triggered:**")
            st.write(ev["how"])
            st.write("**dataLayer:**")
            st.json(ev["datalayer"])
