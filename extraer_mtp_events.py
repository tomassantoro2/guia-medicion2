"""
Extract MTP (Master Tagging Plan) Events sheet: event_name and dataLayer objects → JSON.

Columns referenced: E (script), F (variable name), G (value) — same layout as the Beautify PE template.
"""
import argparse
import json
import os

import openpyxl

_DEFAULT_EXCEL = os.path.join(
    os.path.dirname(__file__), "MTP - Beautify PE - GA4 6jun2023 1.xlsx"
)
_DEFAULT_OUT = os.path.join(os.path.dirname(__file__), "catalogs", "mtp_events.json")


def extract_mtp_events(excel_path: str, sheet_name: str = "Events"):
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        rows = list(
            ws.iter_rows(min_row=1, max_row=600, min_col=1, max_col=10, values_only=True)
        )
    finally:
        wb.close()

    starts = []
    for i, row in enumerate(rows):
        if row[4] and "dataLayer.push" in str(row[4]):
            starts.append(i)

    events_mtp = []
    for start_idx in starts:
        desc = ""
        for j in range(start_idx - 1, max(0, start_idx - 50), -1):
            if j < len(rows) and rows[j][4]:
                val = str(rows[j][4]).strip()
                if val.startswith("Event :") or (val.startswith("Event:") and len(val) > 10):
                    desc = val
                    break
                if "Script" in val and "Variable" in val:
                    break
        dl = {}
        for k in range(start_idx + 1, min(start_idx + 25, len(rows))):
            row = rows[k]
            script_line = row[4]
            if script_line and "});" in str(script_line):
                break
            var, val = row[5], row[6]
            if var and str(var).strip():
                key = str(var).strip()
                dl[key] = str(val).strip() if val is not None else ""
        event_name = dl.get("event_name", dl.get("event", ""))
        if not event_name and not dl:
            continue
        events_mtp.append(
            {
                "event_name": event_name or desc[:50],
                "description": desc,
                "dl": dl,
            }
        )

    return events_mtp


def main():
    parser = argparse.ArgumentParser(description="Extract MTP Events → JSON for the measurement guide.")
    parser.add_argument(
        "--excel",
        default=_DEFAULT_EXCEL,
        help="Path to the MTP Excel file.",
    )
    parser.add_argument(
        "--sheet",
        default="Events",
        help="Worksheet name containing the Events table.",
    )
    parser.add_argument(
        "--output",
        "-o",
        default=_DEFAULT_OUT,
        help="Output JSON path (default: catalogs/mtp_events.json).",
    )
    args = parser.parse_args()

    if not os.path.isfile(args.excel):
        raise SystemExit(f"Excel file not found: {args.excel}")

    events = extract_mtp_events(args.excel, args.sheet)
    out_dir = os.path.dirname(os.path.abspath(args.output))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(events, f, indent=2, ensure_ascii=False)
    print(f"Saved {len(events)} events to {args.output}")


if __name__ == "__main__":
    main()
